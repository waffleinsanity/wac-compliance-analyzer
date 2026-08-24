"""Evidence Log.xlsx fill + stable exhibit ordinals for IR/SOD superscript cites.

Template: data/templates/Evidence Log.xlsx (official DOH evidence log shell).
Export prefers investigator-edited ``report.evidence_log`` when present; otherwise
rows are built from CaseEvidence ordered by created_at (then id).
"""

from __future__ import annotations

import io
import json
import re
from copy import copy
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.config import settings
from app.database import CaseEvidence, InvestigationCase, User
from app.schemas import EvidenceLogDraft, EvidenceLogRow, InvestigationReport

BLANK_FILENAME = "Evidence Log.xlsx"
_DATA_START_ROW = 4
_MAX_WAC_COLS = 4  # H–K

_SUPERSCRIPT_MAP = str.maketrans("0123456789", "⁰¹²³⁴⁵⁶⁷⁸⁹")


@dataclass(frozen=True)
class ExhibitRow:
    exhibit_no: int
    evidence: CaseEvidence

    @property
    def evidence_id(self) -> int:
        return int(self.evidence.id)

    @property
    def label(self) -> str:
        return f"#{self.exhibit_no}"

    @property
    def superscript(self) -> str:
        return exhibit_superscript(self.exhibit_no)


def blank_evidence_log_path() -> Path:
    primary = settings.templates_dir / BLANK_FILENAME
    if primary.is_file():
        return primary
    return settings.examples_dir / "policy_guidance" / BLANK_FILENAME


def exhibit_superscript(n: int) -> str:
    """Unicode superscript digits for plain-text IR/SOD lines."""
    if n <= 0:
        return ""
    return str(int(n)).translate(_SUPERSCRIPT_MAP)


def list_exhibits_for_case(
    case: InvestigationCase | None = None,
    *,
    evidence_rows: list[CaseEvidence] | None = None,
) -> list[ExhibitRow]:
    """Stable exhibit numbers: upload order by created_at, then id."""
    if evidence_rows is not None:
        rows = list(evidence_rows)
    elif case is not None:
        rows = list(case.evidence or [])
    else:
        rows = []
    rows.sort(key=lambda e: (e.created_at or datetime.min, e.id or 0))
    return [ExhibitRow(exhibit_no=i, evidence=ev) for i, ev in enumerate(rows, start=1)]


def exhibit_map_by_id(exhibits: list[ExhibitRow]) -> dict[int, ExhibitRow]:
    return {ex.evidence_id: ex for ex in exhibits}


def exhibit_map_by_title(exhibits: list[ExhibitRow]) -> dict[str, ExhibitRow]:
    out: dict[str, ExhibitRow] = {}
    for ex in exhibits:
        title = (ex.evidence.title or ex.evidence.original_filename or "").strip().lower()
        if title and title not in out:
            out[title] = ex
        stem = Path(ex.evidence.original_filename or "").stem.strip().lower()
        if stem and stem not in out:
            out[stem] = ex
        # Display title without extension
        display = re.sub(r"\.(pdf|docx?|txt|md|png|jpe?g|webp)$", "", title, flags=re.I).strip()
        if display and display not in out:
            out[display] = ex
    return out


def _user_display_name(user: User | None) -> str:
    if user is None:
        return ""
    return (getattr(user, "display_name", None) or user.username or getattr(user, "email", "") or "").strip()


def _parse_linked_wacs(raw: str | list[str] | None) -> list[str]:
    if isinstance(raw, list):
        return [str(x).strip() for x in raw if str(x).strip()]
    text = (raw or "").strip()
    if not text:
        return []
    try:
        data = json.loads(text)
        if isinstance(data, list):
            return [str(x).strip() for x in data if str(x).strip()]
    except Exception:
        pass
    return [s.strip() for s in text.split(",") if s.strip()]


def _format_wac_cell(code: str) -> str:
    """Normalize WA WAC/RCW cells only; leave foreign or free-text cites unchanged."""
    text = (code or "").strip()
    if not text:
        return ""
    upper = text.upper()
    if upper.startswith("WAC ") or upper.startswith("RCW "):
        return text
    bare = re.sub(r"(?i)^(WAC|RCW)\s+", "", text).strip()
    if not bare:
        return ""
    if re.match(r"(?i)^rcw\b", bare) or bare.upper().startswith("RCW"):
        return bare if bare.upper().startswith("RCW") else f"RCW {bare}"
    # Washington behavioral-health chapter only; do not invent WAC authority.
    if re.match(r"^246-\d+", bare):
        return f"WAC {bare}"
    return text


def _facility_name(report: InvestigationReport | dict[str, Any] | None, case: InvestigationCase) -> str:
    data: dict[str, Any] = {}
    if isinstance(report, InvestigationReport):
        data = report.model_dump()
    elif isinstance(report, dict):
        data = report
    sod = data.get("sod") or {}
    fi = data.get("facility_info") or {}
    name = (sod.get("facility_name") or "").strip()
    if name:
        return name
    addr = (sod.get("facility_address") or fi.get("facility_address") or case.facility_address or "").strip()
    if addr:
        return addr.split("\n", 1)[0].strip()
    return (case.title or "").strip()


def _license_number(report: InvestigationReport | dict[str, Any] | None, case: InvestigationCase) -> str:
    data: dict[str, Any] = {}
    if isinstance(report, InvestigationReport):
        data = report.model_dump()
    elif isinstance(report, dict):
        data = report
    sod = data.get("sod") or {}
    fi = data.get("facility_info") or {}
    return (
        (sod.get("credential_number") or fi.get("credential_number") or case.credential_number or "")
        .strip()
    )


def _case_number(report: InvestigationReport | dict[str, Any] | None, case: InvestigationCase) -> str:
    data: dict[str, Any] = {}
    if isinstance(report, InvestigationReport):
        data = report.model_dump()
    elif isinstance(report, dict):
        data = report
    return (case.case_id_label or data.get("case_id") or str(case.id)).strip()


def _copy_row_style(ws, src_row: int, dest_row: int) -> None:
    for col in range(1, (ws.max_column or 11) + 1):
        src = ws.cell(src_row, col)
        dest = ws.cell(dest_row, col)
        if src.has_style:
            dest._style = copy(src._style)
        dest.number_format = src.number_format


def _clear_data_row(ws, row: int) -> None:
    for col in range(1, (ws.max_column or 11) + 1):
        ws.cell(row, col).value = None


def _date_collected(ev: CaseEvidence) -> str:
    dt = ev.created_at
    if not dt:
        return ""
    if dt.tzinfo is not None:
        dt = dt.replace(tzinfo=None)
    return dt.strftime("%m-%d-%y")


def _collector_name(db: Session | None, ev: CaseEvidence, fallback_user: User | None) -> str:
    if db is not None and ev.uploaded_by:
        uploader = db.get(User, ev.uploaded_by)
        name = _user_display_name(uploader)
        if name:
            return name
    return _user_display_name(fallback_user)


def draft_evidence_log_from_case(
    *,
    case: InvestigationCase,
    report: InvestigationReport | dict[str, Any] | None,
    user: User | None,
    exhibits: list[ExhibitRow] | None = None,
    db: Session | None = None,
) -> EvidenceLogDraft:
    """Build an editable Evidence Log draft from case uploads + header fields."""
    rows_src = exhibits if exhibits is not None else list_exhibits_for_case(case)
    rows: list[EvidenceLogRow] = []
    for ex in rows_src:
        ev = ex.evidence
        links = _parse_linked_wacs(ev.linked_wac_ids)
        rows.append(
            EvidenceLogRow(
                exhibit_number=ex.exhibit_no,
                description=(ev.title or ev.original_filename or f"document {ev.id}").strip(),
                date_collected=_date_collected(ev),
                collected_by=_collector_name(db, ev, user),
                method="Electronic upload",
                electronic_location=(ev.stored_path or "").replace("\\", "/"),
                wac_codes=[_format_wac_cell(c) for c in links[:_MAX_WAC_COLS]],
                evidence_id=ex.evidence_id,
            )
        )
    return EvidenceLogDraft(
        investigator_name=_user_display_name(user),
        case_numbers=_case_number(report, case),
        license_numbers=_license_number(report, case),
        facility_name=_facility_name(report, case),
        rows=rows,
    )


def evidence_log_from_report(
    report: InvestigationReport | dict[str, Any] | None,
) -> EvidenceLogDraft | None:
    if report is None:
        return None
    data = report.model_dump() if isinstance(report, InvestigationReport) else dict(report)
    raw = data.get("evidence_log")
    if not raw:
        return None
    try:
        return EvidenceLogDraft.model_validate(raw)
    except Exception:
        return None


def build_evidence_log_xlsx(
    *,
    case: InvestigationCase,
    report: InvestigationReport | dict[str, Any] | None,
    user: User | None,
    exhibits: list[ExhibitRow],
    db: Session | None = None,
    draft: EvidenceLogDraft | None = None,
) -> bytes:
    """Fill Investigation Evidence Log template; return .xlsx bytes.

    Prefer investigator-edited ``draft`` (or ``report.evidence_log``) when present.
    """
    path = blank_evidence_log_path()
    if not path.is_file():
        raise FileNotFoundError(f"Evidence Log template missing: {path}")

    log = draft or evidence_log_from_report(report)
    if log is None:
        log = draft_evidence_log_from_case(
            case=case, report=report, user=user, exhibits=exhibits, db=db
        )

    wb = load_workbook(path)
    ws = wb.active

    investigator = (log.investigator_name or _user_display_name(user)).strip()
    ws["A2"] = f"Investigator Name: {investigator}".rstrip()
    ws["B2"] = f"Case Numbers: {(log.case_numbers or _case_number(report, case)).strip()}"
    ws["C2"] = f"License Numbers: {(log.license_numbers or _license_number(report, case)).strip()}"
    ws["D2"] = f"Facility/Agency Name: {(log.facility_name or _facility_name(report, case)).strip()}"

    last_template_row = max(ws.max_row, _DATA_START_ROW)
    data_rows = list(log.rows or [])
    needed = _DATA_START_ROW + max(len(data_rows), 1) - 1
    while ws.max_row < needed:
        new_row = ws.max_row + 1
        _copy_row_style(ws, last_template_row, new_row)

    for i, row_data in enumerate(data_rows):
        row = _DATA_START_ROW + i
        num = int(row_data.exhibit_number or (i + 1))
        ws.cell(row, 1).value = f"#{num}"
        ws.cell(row, 2).value = (row_data.description or "").strip()
        ws.cell(row, 3).value = (row_data.date_collected or "").strip()
        ws.cell(row, 4).value = (row_data.collected_by or "").strip()
        ws.cell(row, 5).value = (row_data.method or "Electronic upload").strip()
        ws.cell(row, 6).value = (row_data.electronic_location or "").replace("\\", "/")
        links = list(row_data.wac_codes or [])
        for wi in range(_MAX_WAC_COLS):
            cell = ws.cell(row, 8 + wi)
            if wi < len(links) and str(links[wi]).strip():
                cell.value = _format_wac_cell(str(links[wi]))
            else:
                cell.value = None

    clear_from = _DATA_START_ROW + len(data_rows)
    for row in range(clear_from, ws.max_row + 1):
        _clear_data_row(ws, row)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def exhibit_number_map_from_draft(draft: EvidenceLogDraft | None) -> dict[int, int]:
    """Map CaseEvidence id → exhibit number from an edited log."""
    out: dict[int, int] = {}
    if not draft:
        return out
    for row in draft.rows or []:
        if row.evidence_id is None:
            continue
        try:
            out[int(row.evidence_id)] = int(row.exhibit_number)
        except (TypeError, ValueError):
            continue
    return out


def apply_draft_exhibit_numbers(
    exhibits: list[ExhibitRow],
    draft: EvidenceLogDraft | None,
) -> list[ExhibitRow]:
    """Prefer investigator-edited exhibit ordinals when the Evidence Log draft maps them."""
    remap = exhibit_number_map_from_draft(draft)
    if not remap:
        return exhibits
    return [
        ExhibitRow(
            exhibit_no=int(remap.get(ex.evidence_id, ex.exhibit_no)),
            evidence=ex.evidence,
        )
        for ex in exhibits
    ]


def exhibits_for_report(
    case: InvestigationCase,
    report: InvestigationReport | dict[str, Any] | None = None,
    *,
    evidence_rows: list[CaseEvidence] | None = None,
) -> list[ExhibitRow]:
    """Upload-order exhibits, remapped by ``report.evidence_log`` when present."""
    base = list_exhibits_for_case(case, evidence_rows=evidence_rows)
    return apply_draft_exhibit_numbers(base, evidence_log_from_report(report))


def append_exhibit_superscript(text: str, exhibit_no: int) -> str:
    """Append Unicode superscript after text if not already present."""
    mark = exhibit_superscript(exhibit_no)
    if not mark:
        return text or ""
    body = (text or "").rstrip()
    if body.endswith(mark):
        return body
    if body and body[-1] in "⁰¹²³⁴⁵⁶⁷⁸⁹":
        return body
    return f"{body}{mark}"


def append_exhibit_superscripts(text: str, exhibit_nos: list[int]) -> str:
    body = (text or "").rstrip()
    for n in exhibit_nos:
        body = append_exhibit_superscript(body, n)
    return body


def resolve_exhibit_nos_for_finding(
    finding: dict[str, Any] | Any,
    by_id: dict[int, ExhibitRow],
) -> list[int]:
    if hasattr(finding, "model_dump"):
        d = finding.model_dump()
    else:
        d = dict(finding or {})
    nos: list[int] = []
    for raw in d.get("evidence_ids") or []:
        try:
            eid = int(str(raw).strip())
        except ValueError:
            continue
        ex = by_id.get(eid)
        if ex:
            nos.append(ex.exhibit_no)
    return nos


def match_exhibit_for_review_line(line: str, by_title: dict[str, ExhibitRow]) -> ExhibitRow | None:
    """Match a Document Review line to an exhibit by quoted title."""
    m = re.match(
        r'^The investigator reviewed ["“](.+?)["”] dated ',
        (line or "").strip(),
        flags=re.I,
    )
    if not m:
        return None
    title = m.group(1).strip().lower()
    key = re.sub(r"\.(pdf|docx?|txt|md|png|jpe?g|webp)$", "", title, flags=re.I).strip()
    return by_title.get(title) or by_title.get(key)


_TRAILING_SUPER_RE = re.compile(r"[⁰¹²³⁴⁵⁶⁷⁸⁹]+$")


def strip_trailing_superscripts(text: str) -> tuple[str, str]:
    """Split body and trailing Unicode superscript digits."""
    body = (text or "").rstrip()
    m = _TRAILING_SUPER_RE.search(body)
    if not m:
        return body, ""
    return body[: m.start()].rstrip(), m.group(0)


def annotate_process_with_exhibits(
    process: list[str],
    exhibits: list[ExhibitRow],
) -> list[str]:
    """Append Unicode exhibit superscripts to Document Review lines."""
    by_title = exhibit_map_by_title(exhibits)
    out: list[str] = []
    for line in process or []:
        base, _ = strip_trailing_superscripts(line)
        ex = match_exhibit_for_review_line(base, by_title)
        if ex:
            out.append(append_exhibit_superscript(base, ex.exhibit_no))
        else:
            out.append(line)
    return out


def annotate_finding_text_with_exhibits(
    text: str,
    evidence_ids: list[str] | list[int] | None,
    by_id: dict[int, ExhibitRow],
) -> str:
    nos: list[int] = []
    for raw in evidence_ids or []:
        try:
            eid = int(str(raw).strip())
        except ValueError:
            continue
        ex = by_id.get(eid)
        if ex:
            nos.append(ex.exhibit_no)
    if not nos:
        return text or ""
    base, _ = strip_trailing_superscripts(text or "")
    return append_exhibit_superscripts(base, nos)
