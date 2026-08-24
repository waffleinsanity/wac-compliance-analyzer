"""Evidence Log.xlsx fill, pack membership, and IR/SOD superscript cites."""

from __future__ import annotations

import io
import zipfile
from datetime import datetime, timezone

from openpyxl import load_workbook

from app.database import CaseEvidence, InvestigationCase, User
from app.schemas import (
    FacilityInfo,
    InvestigationReport,
    SodDeficiency,
    SodFinding,
    StatementOfDeficiency,
)
from app.services.case_store import dumps_list
from app.services.docx_export import build_investigation_docx, build_sod_docx
from app.services.evidence_log import (
    annotate_process_with_exhibits,
    build_evidence_log_xlsx,
    exhibit_superscript,
    list_exhibits_for_case,
)


def _sample_report(**overrides) -> InvestigationReport:
    base = dict(
        title="Investigative Report",
        subtitle="",
        investigation_date="2026-08-01",
        case_id="EL-2026-1",
        facility_info=FacilityInfo(
            facility_address="123 Facility Rd",
            credential_number="RTF.FS.60855826",
        ),
        intake_details="Intake.",
        allegation_preamble="",
        allegations=[],
        investigative_process=[
            "Document Review",
            'The investigator reviewed "Policy" dated August 1, 2026.',
        ],
        summary_of_findings="",
        conclusions=[],
        actions="",
        comparisons=[],
        findings=[],
        report_text="",
        selected_count=0,
        duration_ms=0,
        document_preview="",
        sod=StatementOfDeficiency(
            title="Statement of Deficiencies",
            facility_name="Test Agency",
            facility_address="123 Facility Rd",
            case_id="EL-2026-1",
            credential_number="RTF.FS.60855826",
            deficiencies=[
                SodDeficiency(
                    id="d1",
                    regulation_cite="WAC 246-341-0600(1)",
                    regulation_text="Store text.",
                    based_on="Based on record review.",
                    failure_to="Failure to comply.",
                    findings=[],
                )
            ],
        ),
    )
    base.update(overrides)
    return InvestigationReport(**base)


def _case_with_evidence(db, user: User, *, title: str = "Policy.pdf") -> InvestigationCase:
    report = _sample_report()
    case = InvestigationCase(
        owner_user_id=user.id,
        case_id_label="EL-2026-1",
        title="Evidence log case",
        status="draft",
        complaint_text="Complaint text for evidence log tests.",
        facility_address="123 Facility Rd\nOlympia, WA",
        credential_number="RTF.FS.60855826",
        approved_wac_ids=dumps_list(["WAC 246-341-0600"]),
        status_changed_by=user.id,
        current_report_json=report.model_dump_json(),
    )
    db.add(case)
    db.commit()
    db.refresh(case)
    ev = CaseEvidence(
        case_id=case.id,
        title=title,
        original_filename=title,
        stored_path=f"{case.id}/evidence/test_{title}",
        content_type="application/pdf",
        linked_wac_ids=dumps_list(["246-341-0600"]),
        notes="",
        uploaded_by=user.id,
        created_at=datetime(2026, 8, 1, 12, 0, 0, tzinfo=timezone.utc),
    )
    db.add(ev)
    db.commit()
    db.refresh(case)
    return case


def test_exhibit_superscript_digits():
    assert exhibit_superscript(1) == "¹"
    assert exhibit_superscript(12) == "¹²"


def test_build_evidence_log_xlsx_header_and_row(db, auth_user):
    case = _case_with_evidence(db, auth_user)
    exhibits = list_exhibits_for_case(case)
    assert len(exhibits) == 1
    assert exhibits[0].exhibit_no == 1

    raw = build_evidence_log_xlsx(
        case=case,
        report=_sample_report(),
        user=auth_user,
        exhibits=exhibits,
        db=db,
    )
    wb = load_workbook(io.BytesIO(raw))
    ws = wb.active
    assert "Investigator Name:" in str(ws["A2"].value or "")
    assert "EL-2026-1" in str(ws["B2"].value or "")
    assert "RTF.FS.60855826" in str(ws["C2"].value or "")
    assert ws["A4"].value == "#1"
    assert "Policy" in str(ws["B4"].value or "")
    assert ws["E4"].value == "Electronic upload"
    assert ws["A5"].value in (None, "")


def test_export_pack_includes_evidence_log(client, db, auth_user):
    case = _case_with_evidence(db, auth_user)
    res = client.post(f"/api/cases/{case.id}/export/pack?acknowledge_gaps=true")
    assert res.status_code == 200, res.text
    with zipfile.ZipFile(io.BytesIO(res.content)) as zf:
        names = zf.namelist()
    assert any(n.startswith("Evidence_Log_") and n.endswith(".xlsx") for n in names)
    assert any(n.startswith("IR_") and n.endswith(".docx") for n in names)
    assert any(n.startswith("SOD_") and n.endswith(".docx") for n in names)


def test_export_evidence_log_standalone(client, db, auth_user):
    case = _case_with_evidence(db, auth_user)
    res = client.post(f"/api/cases/{case.id}/export/evidence-log")
    assert res.status_code == 200, res.text
    assert "spreadsheetml" in (res.headers.get("content-type") or "")
    wb = load_workbook(io.BytesIO(res.content))
    assert wb.active["A4"].value == "#1"


def test_export_evidence_log_empty_case_ok(client, db, auth_user):
    report = _sample_report()
    case = InvestigationCase(
        owner_user_id=auth_user.id,
        case_id_label="EL-EMPTY",
        title="Empty evidence",
        status="draft",
        complaint_text="Complaint.",
        approved_wac_ids=dumps_list([]),
        status_changed_by=auth_user.id,
        current_report_json=report.model_dump_json(),
    )
    db.add(case)
    db.commit()
    db.refresh(case)
    res = client.post(f"/api/cases/{case.id}/export/evidence-log")
    assert res.status_code == 200, res.text
    wb = load_workbook(io.BytesIO(res.content))
    assert "EL-EMPTY" in str(wb.active["B2"].value or "")
    assert wb.active["A4"].value in (None, "")


def test_export_evidence_log_blocked_when_trashed(client, db, auth_user):
    case = _case_with_evidence(db, auth_user)
    case.status = "trashed"
    db.add(case)
    db.commit()
    res = client.post(f"/api/cases/{case.id}/export/evidence-log")
    assert res.status_code == 400, res.text
    assert "trash" in (res.json().get("detail") or "").lower()


def test_export_evidence_log_unknown_case_404(client, auth_user):
    res = client.post("/api/cases/999999/export/evidence-log")
    assert res.status_code == 404


def test_draft_exhibit_numbers_remap_for_detail(client, db, auth_user):
    from app.schemas import EvidenceLogDraft, EvidenceLogRow

    case = _case_with_evidence(db, auth_user)
    ev_id = case.evidence[0].id
    report = _sample_report(
        evidence_log=EvidenceLogDraft(
            investigator_name="Tester",
            case_numbers="EL-2026-1",
            license_numbers="RTF.FS.60855826",
            facility_name="Test Agency",
            rows=[
                EvidenceLogRow(
                    exhibit_number=7,
                    description="Policy",
                    date_collected="08-01-26",
                    collected_by="Tester",
                    method="Electronic upload",
                    electronic_location="x",
                    wac_codes=["WAC 246-341-0600"],
                    evidence_id=ev_id,
                )
            ],
        )
    )
    case.current_report_json = report.model_dump_json()
    db.add(case)
    db.commit()

    res = client.get(f"/api/cases/{case.id}")
    assert res.status_code == 200, res.text
    assert res.json()["evidence"][0]["exhibit_number"] == 7

    xlsx = client.post(f"/api/cases/{case.id}/export/evidence-log")
    assert xlsx.status_code == 200, xlsx.text
    wb = load_workbook(io.BytesIO(xlsx.content))
    assert wb.active["A4"].value == "#7"


def test_format_wac_cell_skips_foreign_cites():
    from app.services.evidence_log import _format_wac_cell

    assert _format_wac_cell("246-341-0600") == "WAC 246-341-0600"
    assert _format_wac_cell("WAC 246-341-0600") == "WAC 246-341-0600"
    assert _format_wac_cell("Title 22 CCR § 70577") == "Title 22 CCR § 70577"
    assert _format_wac_cell("OR 309-019-0110") == "OR 309-019-0110"


def test_evidence_out_includes_exhibit_number(client, db, auth_user):
    case = _case_with_evidence(db, auth_user)
    res = client.get(f"/api/cases/{case.id}")
    assert res.status_code == 200, res.text
    evidence = res.json().get("evidence") or []
    assert len(evidence) == 1
    assert evidence[0].get("exhibit_number") == 1


def _docx_xml(raw: bytes) -> str:
    with zipfile.ZipFile(io.BytesIO(raw)) as zf:
        return zf.read("word/document.xml").decode("utf-8")


def test_ir_docx_has_superscript_vert_align(db, auth_user):
    case = _case_with_evidence(db, auth_user)
    exhibits = list_exhibits_for_case(case)
    process = annotate_process_with_exhibits(
        [
            "Document Review",
            'The investigator reviewed "Policy" dated August 1, 2026.',
        ],
        exhibits,
    )
    assert process[1].endswith("¹")
    report = _sample_report(investigative_process=process)
    raw = build_investigation_docx(report, exhibits=exhibits)
    xml = _docx_xml(raw)
    assert 'w:val="superscript"' in xml


def test_sod_docx_has_superscript_when_evidence_linked(db, auth_user):
    case = _case_with_evidence(db, auth_user)
    exhibits = list_exhibits_for_case(case)
    eid = str(exhibits[0].evidence_id)
    sod = StatementOfDeficiency(
        title="Statement of Deficiencies",
        facility_name="Test Agency",
        facility_address="123 Facility Rd",
        case_id="EL-2026-1",
        credential_number="RTF.FS.60855826",
        deficiencies=[
            SodDeficiency(
                id="d1",
                regulation_cite="WAC 246-341-0600(1)",
                regulation_text="Store text.",
                based_on="Based on record review.",
                failure_to="Failure to comply.",
                findings=[
                    SodFinding(
                        method="document review",
                        text="Reviewed facility policy.",
                        evidence_ids=[eid],
                    )
                ],
            )
        ],
    )
    report = _sample_report(sod=sod)
    raw = build_sod_docx(report, exhibits=exhibits)
    xml = _docx_xml(raw)
    assert 'w:val="superscript"' in xml
